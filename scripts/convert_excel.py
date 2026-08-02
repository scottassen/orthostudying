#!/usr/bin/env python3
"""Convert the National Consensus recall workbook into normalized JSON.

Outputs:
  data/questions.json         - every substantive question row, all sheets
  data/questions_recent.json  - priority set: 2018-2025 sheets + Controversial

Each record:
  {
    "id": "<sheet>#<row>",
    "year": 2024,                # inferred exam year (int) or None
    "sheet": "2024-Fall (Consensus)",
    "q_number": "12",            # question number if present
    "question": "...",           # question text WITH options stripped where possible
    "options": "...",            # raw option text if in a separate column
    "consensus_answer": "B",
    "evidence": "...",           # evidence / explanation / notes column text
    "topic_hint": "...",         # explicit topic column (Controversial sheet only)
    "repeat_of": "...",          # 'Type'/years-repeated column text if present
    "has_image": false,          # row overlaps an embedded image anchor
    "row": 3
  }

Run:  python3 scripts/convert_excel.py
"""
import json
import re
import os
import sys

import openpyxl

XLSX = "2025 National Consensus-3.xlsx"
OUT_DIR = "data"

# Sheets whose exam year is the sheet name (or embedded in it).
YEAR_RE = re.compile(r"(20\d{2})")

# Header keywords -> canonical field. Matched case-insensitively as substrings.
HEADER_MAP = {
    "question": ["question + answer", "question", "mcqs"],
    "options": ["options"],
    "consensus_answer": ["2025 consensus", "consensus", "answer"],
    "evidence": ["evidence", "explanation", "notes", "discussion"],
    "topic_hint": ["topic"],
    "q_number": ["question \n#", "question #", "#"],
    "repeat_of": ["type (new", "prior consensus"],
}


def norm(s):
    if s is None:
        return None
    s = str(s).strip()
    return s if s else None


def find_header_row(ws, max_scan=6):
    """Return (row_index, {col_index: header_text}) for the most header-like row."""
    best = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        labels = {j: norm(c) for j, c in enumerate(row) if norm(c)}
        # A header row has several short-ish text labels including a known keyword.
        score = 0
        for txt in labels.values():
            t = txt.lower()
            if any(k in t for kws in HEADER_MAP.values() for k in kws):
                score += 1
        if best is None or score > best[2]:
            best = (i, labels, score)
    return best[0], best[1]


# Fields resolved in this priority order; once a column is claimed it is not
# reused. This prevents e.g. the "answer" keyword from claiming a
# "Question + answer choices" column that belongs to `question`.
FIELD_ORDER = ["q_number", "topic_hint", "question", "options",
               "consensus_answer", "evidence", "repeat_of"]


def map_columns(headers):
    """Given {col: header_text} map canonical field -> col index (claims once)."""
    col = {}
    claimed = set()
    for field in FIELD_ORDER:
        for kw in HEADER_MAP[field]:
            for c, txt in sorted(headers.items()):
                if c in claimed:
                    continue
                if kw in txt.lower():
                    col[field] = c
                    claimed.add(c)
                    break
            if field in col:
                break
    return col


# Answer-letter prefix like "a)" "A." "d) cane" at start of an option line.
def strip_options_from_question(text):
    """Split a combined 'stem a) .. b) ..' cell into (stem, options)."""
    if not text:
        return text, None
    # Find first option marker at line start or after newline.
    m = re.search(r"(?:^|\n)\s*[a-eA-E][\).]\s", text)
    if not m:
        return text, None
    stem = text[: m.start()].strip()
    opts = text[m.start():].strip()
    if len(stem) < 10:  # marker was spurious; keep whole thing as question
        return text, None
    return stem, opts


def sheet_year(name):
    m = YEAR_RE.search(name)
    return int(m.group(1)) if m else None


RECENT_SHEETS_MIN_YEAR = 2018
SPECIAL_RECENT = {"Controversial Questions", "Fall 2025 (Non Consensus)",
                  "2024-Fall (Consensus)", "2024-Spring", "Wrong Anki"}


def is_recent(name):
    y = sheet_year(name)
    if name in SPECIAL_RECENT:
        return True
    return y is not None and y >= RECENT_SHEETS_MIN_YEAR


def image_rows_for_sheet(sheet_index):
    """Return set of row numbers (1-based) that anchor an embedded image.

    Parsed from xl/drawings via the sheet->drawing rels. Best-effort.
    """
    import zipfile
    import xml.etree.ElementTree as ET
    rows = set()
    try:
        z = zipfile.ZipFile(XLSX)
        rel_path = f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"
        if rel_path not in z.namelist():
            return rows
        rel_xml = ET.fromstring(z.read(rel_path))
        ns_r = "{http://schemas.openxmlformats.org/package/2006/relationships}"
        draw_targets = [r.get("Target") for r in rel_xml
                        if "drawing" in (r.get("Target") or "")]
        for t in draw_targets:
            dpath = "xl/" + t.replace("../", "")
            if dpath not in z.namelist():
                continue
            dxml = ET.fromstring(z.read(dpath))
            ns_xdr = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"
            for anchor in dxml.iter():
                if anchor.tag.endswith("}from"):
                    r = anchor.find(f"{ns_xdr}row")
                    if r is not None and r.text is not None:
                        rows.add(int(r.text) + 1)  # xdr rows are 0-based
    except Exception as e:
        print(f"  (image-row parse failed for sheet{sheet_index}: {e})", file=sys.stderr)
    return rows


def parse_workbook():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    all_records = []
    # Map worksheet title -> 1-based sheet index (order in workbook.xml).
    sheet_index = {ws.title: i + 1 for i, ws in enumerate(wb.worksheets)}

    for ws in wb.worksheets:
        name = ws.title
        img_rows = image_rows_for_sheet(sheet_index[name])
        year = sheet_year(name)

        # Wrong Anki has an irregular 3-col layout with no header.
        if name == "Wrong Anki":
            for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                years, q, ans = (list(row) + [None, None, None])[:3]
                q = norm(q)
                if not q or len(q) < 20:
                    continue
                stem, opts = strip_options_from_question(q)
                all_records.append({
                    "id": f"{name}#{ridx}", "year": year, "sheet": name,
                    "q_number": None, "question": stem, "options": opts,
                    "consensus_answer": norm(ans), "evidence": None,
                    "topic_hint": None, "repeat_of": norm(years),
                    "has_image": ridx in img_rows, "row": ridx,
                })
            continue

        hdr_row, headers = find_header_row(ws)
        col = map_columns(headers)
        # Fallback: if no question column detected, use first column.
        qcol = col.get("question", 0)

        for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if ridx <= hdr_row + 1:  # skip header + any pre-header rows
                continue
            row = list(row)

            def get(field):
                c = col.get(field)
                if c is None or c >= len(row):
                    return None
                return norm(row[c])

            qtext = norm(row[qcol]) if qcol < len(row) else None
            if not qtext or len(qtext) < 20:
                continue  # not a substantive question row

            options = get("options")
            if options is None:
                # combined stem+options cell
                qtext, options = strip_options_from_question(qtext)

            all_records.append({
                "id": f"{name}#{ridx}",
                "year": year,
                "sheet": name,
                "q_number": get("q_number"),
                "question": qtext,
                "options": options,
                "consensus_answer": get("consensus_answer"),
                "evidence": get("evidence"),
                "topic_hint": get("topic_hint"),
                "repeat_of": get("repeat_of"),
                "has_image": ridx in img_rows,
                "row": ridx,
            })

    return all_records


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    records = parse_workbook()
    with open(os.path.join(OUT_DIR, "questions.json"), "w") as f:
        json.dump(records, f, ensure_ascii=False, indent=1)
    recent = [r for r in records if is_recent(r["sheet"])]
    with open(os.path.join(OUT_DIR, "questions_recent.json"), "w") as f:
        json.dump(recent, f, ensure_ascii=False, indent=1)

    print(f"total records: {len(records)}")
    print(f"recent records: {len(recent)}")
    by_sheet = {}
    for r in records:
        by_sheet[r["sheet"]] = by_sheet.get(r["sheet"], 0) + 1
    for s, n in by_sheet.items():
        print(f"  {s}: {n}")
    print(f"with image anchor: {sum(1 for r in records if r['has_image'])}")


if __name__ == "__main__":
    main()
